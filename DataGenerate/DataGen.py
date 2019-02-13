import openpyxl


def data_generator():
    vk = openpyxl.load_workbook('C:\\Users\\hsouza\\Projects\\E2EAutomation\\Data.xlsx')
    sh = vk['Folha1']
    r = sh.max_row
    li = []
    li1 = []
    for i in range(1, r+1):
        li1 = []
        un = sh.cell(i, 1)
        up = sh.cell(i, 2)
        li1.insert(0, un.value)
        li1.insert(1, up.value)
        li.insert(i-1, li1)
    print(li)
    return li

    '''
    li = [['uname1', 'pass1'], ['uname2', 'pass2'], ['uname3', 'pass3']]
    return li
    '''